import os
import asyncio
import json
import re
from enum import Enum
from typing import List, Tuple, Dict
from pydantic import BaseModel, ValidationError
from app.config.bedrock_llm import llm_config
from app.utils import logger
from app.utils.file_utils import get_code_files, get_file_language
import aiofiles
import tiktoken
from concurrent.futures import ThreadPoolExecutor
from functools import lru_cache, wraps
from threading import Lock
import hashlib
import time
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from .vector_store_service import get_vector_store_service
import logging
from app.utils.angular_parser import summarize_angular_file
from app.utils.dotnet_parser import summarize_csharp_code
import inspect

# Constants
CHUNK_GROUP_SIZE = 40
LLM_CONCURRENCY_LIMIT = 10
llm_semaphore = asyncio.Semaphore(LLM_CONCURRENCY_LIMIT)
LINTER_CONCURRENCY = 300
linter_sema = asyncio.Semaphore(LINTER_CONCURRENCY)
MAX_MODEL_TOKENS = 128000
MAX_SAFE_TOKENS = 50000
MAX_EMBEDDING_TOKENS = 8192
BATCH_SIZE = 50
MAX_LINES_PER_CHUNK = 2000
MAX_SECTION_TOKENS = 1000
log_lock = Lock()
vector_store = get_vector_store_service()
VECTOR_STORE_DIR = "vector_store"

# Tokenizer and Chunking Utilities
@lru_cache(maxsize=10)
def get_tokenizer(model: str = "text-embedding-ada-002"):
    return tiktoken.encoding_for_model(model)

def count_tokens(text: str, model: str = "text-embedding-ada-002") -> int:
    tokenizer = get_tokenizer(model)
    return len(tokenizer.encode(text))

def chunk_text_for_embedding(text: str, model: str = "text-embedding-ada-002", max_tokens: int = MAX_EMBEDDING_TOKENS) -> List[str]:
    tokenizer = get_tokenizer(model)
    tokens = tokenizer.encode(text)

    chunks = []
    for i in range(0, len(tokens), max_tokens):
        chunk = tokens[i:i + max_tokens]
        chunks.append(tokenizer.decode(chunk))

    return chunks

# Enhanced configuration for dynamic business rule detection (updated for VBA)
RULES_CONFIG = {
    "financial": {
        "patterns": [
            r"\b(interest_rate|credit_score|risk_score|loan_eligibility|payment_schedule)\s*[+\-*/=]\s*[\d+\.*%*].*",  # Specific financial calculations
            r"\bif.*\b(credit_score|income|debt|risk|loan_amount)\b.*[<>=].*\b(approve|reject|calculate|assess)\b.*",  # Financial decision logic
            r"\bdef\s+(calculate_|compute_|assess_)(interest|loan|risk|score)\b.*",  # Financial calculation functions
            r"\bSub\s+(calculate_|compute_|assess_)(cost|total|adjustment|summary)\b.*",  # VBA financial subs
        ],
        "description": "Proprietary financial calculations, risk assessments, or loan eligibility logic."
    },
    "ecommerce": {
        "patterns": [
            r"\b(discount|price|total_cost|tax|shipping)\s*[+\-*/=]\s*.*(if|where|category|location|amount).*",  # Pricing and discount rules
            r"\bif.*\b(location|category|product_type|order_value|customer_type)\b.*[<>=].*\b(discount|price|promotion)\b.*",  # Conditional pricing logic
            r"\bdef\s+(apply_|calculate_)(discount|promotion|price)\b.*",  # Ecommerce-specific functions
            r"\bSub\s+(apply_|calculate_)(discount|promotion|cost|total)\b.*",  # VBA ecommerce subs
        ],
        "description": "Custom pricing, discount, or promotion rules based on business criteria."
    },
    "general": {
        "patterns": [
            r"\b(custom_|proprietary_|internal_)(rule|logic|calc|workflow)\b.*=\s*.*",  # Custom-named business logic
            r"\bif.*\b(custom_|proprietary_|internal_)\b.*[<>=].*\b(process|validate|compute)\b.*",  # Custom conditional logic
            r"\bdef\s+(custom_|proprietary_|internal_)(process|validate|compute)\b.*",  # Custom processing functions
            r"\bSub\s+(custom_|proprietary_|internal_)(process|validate|compute)\b.*",  # VBA custom subs
        ],
        "description": "Organization-specific business logic, workflows, or proprietary computations."
    },
    "excel": {
        "patterns": [
            r"FORMULA:.*\b(SUM|AVERAGE|IF|VLOOKUP|HLOOKUP|INDEX|MATCH|COUNTIF|SUMIF|ROUND|PMT|FV|PV)\b\s*\(.*\)",  # Excel formulas
            r"FORMULA:.*\b(AND|OR|NOT|IFERROR|IFNA)\b\s*\(.*\)",  # Logical Excel formulas
            r"DATA_VALIDATION:.*\b(list|whole|decimal|textLength|date|time|custom)\b.*",  # Data validation rules
            r"CONDITIONAL_FORMATTING:.*\b(cellIs|expression|colorScale|dataBar|iconSet)\b.*",  # Conditional formatting
            r"FORMULA:.*\[.*\].*",  # Formulas with named ranges or custom references
            r"\bSub\s+.*\b(totalcost|finalcost|tax|discount|overhead)\b.*",  # VBA macros with business logic
        ],
        "description": "Excel-specific business rules including complex formulas, data validation, conditional formatting, and VBA macros."
    }
}

# Model definitions (unchanged)
class GraphResponse(BaseModel):
    target_graph: str
    generated_code: str

class RequirementsResponse(BaseModel):
    requirements: str

class FileDependency(BaseModel):
    file_name: str
    relative_path: str
    dependency_reason: str

class FileRequirements(BaseModel):
    relative_path: str
    file_name: str
    requirements: str
    dependencies: List[FileDependency]
    dependents: List[dict] = []

class FilesRequirements(BaseModel):
    files: List[FileRequirements]
    project_summary: str = ""
    graphs: List[GraphResponse] = []

class OrgSpecificRules(BaseModel):
    relative_path: str
    file_name: str
    rules: str

def parse_vba_macros_from_content(content: str) -> Dict[str, Dict[str, str]]:
    """
    Extract VBA macro names, parameters, and comments from the provided content.
    """
    macros = {}
    lines = content.splitlines()
    current_macro = None
    macro_lines = []
    for line in lines:
        if line.strip().startswith('--- VBA Module:'):
            if current_macro and macro_lines:
                macros[current_macro] = {
                    'parameters': extract_macro_parameters(macro_lines),
                    'comments': extract_macro_comments(macro_lines),
                    'logic': '\n'.join(macro_lines)
                }
            current_macro = line.strip().replace('--- VBA Module:', '').strip()
            macro_lines = []
        elif current_macro is not None:
            macro_lines.append(line)
    if current_macro and macro_lines:
        macros[current_macro] = {
            'parameters': extract_macro_parameters(macro_lines),
            'comments': extract_macro_comments(macro_lines),
            'logic': '\n'.join(macro_lines)
        }
    return macros

def extract_macro_parameters(macro_lines: List[str]) -> str:
    """
    Extract macro parameters from the provided macro lines.
    """
    parameters = []
    for line in macro_lines:
        match = re.search(r'\bSub\s+(\w+)\s*\((.*?)\)', line, re.IGNORECASE)
        if match:
            parameters.append(f"{match.group(1)}({match.group(2)})")
    return ', '.join(parameters)

def extract_macro_comments(macro_lines: List[str]) -> str:
    """
    Extract macro comments from the provided macro lines.
    """
    comments = []
    for line in macro_lines:
        match = re.search(r"'(.*)", line, re.IGNORECASE)
        if match:
            comments.append(match.group(1).strip())
    return '\n'.join(comments)

def summarize_excel_vba_logic(content: str) -> str:
    """
    Parse the extracted Excel/VBA content and produce a detailed, step-by-step summary of all business logic, rules, calculations, and workflow automation as found in the file.
    - For VBA: List each Sub/Function, describe its purpose, logic, and steps in plain business language.
    - For formulas: List each formula, its cell, and what it calculates in business terms.
    - For validations: List all data validation and conditional formatting rules.
    - For named ranges and pivots: List and describe their business meaning.
    """
    lines = content.splitlines()
    summary = []
    macros = parse_vba_macros_from_content(content)
    for macro, details in macros.items():
        summary.append(f"Macro: {macro}")
        summary.append(f"Parameters: {details['parameters']}")
        summary.append(f"Comments: {details['comments']}")
        summary.append(f"Logic: {details['logic']}")
    # Formulas
    for line in lines:
        m = re.match(r'FORMULA: ([A-Z]+\d+) = (.+)', line)
        if m:
            cell, formula = m.groups()
            summary.append(f"Formula in {cell}: {formula}")
    # Data Validations
    for line in lines:
        if line.strip().startswith('DATA_VALIDATION:'):
            summary.append(f"Data Validation Rule: {line.strip().replace('DATA_VALIDATION:', '').strip()}")
    # Conditional Formatting
    for line in lines:
        if line.strip().startswith('CONDITIONAL_FORMATTING:'):
            summary.append(f"Conditional Formatting: {line.strip().replace('CONDITIONAL_FORMATTING:', '').strip()}")
    # Named Ranges
    for line in lines:
        if line.strip().startswith('- ') and ':' in line:
            summary.append(f"Named Range: {line.strip()}")
    # Pivot Tables
    for line in lines:
        if line.strip().startswith('Pivot Tables in') or line.strip().startswith('- Name:'):
            summary.append(line.strip())
    # External Links
    for line in lines:
        if line.strip().startswith('External Link:'):
            summary.append(line.strip())
    # If nothing found, fallback
    if not summary:
        return "No business logic, rules, or automation found in the file."
    return '\n'.join(summary)

async def narrative_business_logic_summary(technical_logic: str, file_path: str) -> dict:
    from app.utils.logger import logger
    logger.info(f"[LLM ENTRY] narrative_business_logic_summary called for file: {file_path}")
    
    # Step 1: Extract all business rules/logic as a list
    rules_prompt = f"""
You are an expert business analyst. Given the following technical summary of an Excel/VBA file, extract ALL business rules, logic, macros, formulas, and workflow steps as a detailed, numbered list. Be exhaustive and do not summarize—list every rule, macro, formula, and workflow step you can find. If there are none, say 'No business rules or logic found.'

TECHNICAL SUMMARY:
{technical_logic}
"""
    max_tokens = 1000
    logger.debug(f"[LLM DEBUG] Rules extraction prompt for {file_path}: {rules_prompt[:1000]}...")
    import asyncio
    loop = asyncio.get_event_loop()
    try:
        rules_response = await loop.run_in_executor(
            None,
            lambda: asyncio.run(llm_config._llm(rules_prompt, max_tokens=max_tokens, temperature=0.2, stop=None))
        )
        logger.debug(f"[LLM DEBUG] Raw LLM rules output for {file_path}: {rules_response[:200]}...")
    except Exception as e:
        logger.error(f"[LLM ERROR] Exception during LLM rules extraction for {file_path}: {e}")
        rules_response = "No business rules or logic found."

    # Step 2: Sectioned requirements prompt, forceful and explicit
    example_output = EXAMPLES['excel']
    sectioned_prompt = f"""
You are a business analyst. Given the following list of business rules, logic, macros, formulas, and workflow steps from an Excel/VBA file, you MUST output ALL SIX SECTIONS: Overview, Objective, Use Case, Key Functionalities, Workflow Summary, Dependent Files. Each section must be non-empty, file-specific, and at least 3 sentences. If you cannot find business logic for a section, explain why, but do not skip the section. If you do not output all six sections, your answer will be rejected. Use the example format below.

Example output:
{example_output}

BUSINESS RULES AND LOGIC:
{rules_response}
"""
    missing_sections = set(REQUIRED_SECTIONS)
    prev_output = ""
    for attempt in range(1, 4):
        if attempt == 1:
            full_prompt = sectioned_prompt
        else:
            full_prompt = f"Your previous output was missing the following sections: {', '.join([s.title() for s in missing_sections])}. You must output ALL SIX SECTIONS, each with at least 3 sentences, using the business rules and logic below. If you do not, your answer will be rejected.\nPrevious output:\n{prev_output}\n\nBUSINESS RULES AND LOGIC:\n{rules_response}\n"
        logger.debug(f"[LLM DEBUG] Sectioned prompt for {file_path} (attempt {attempt}): {full_prompt[:1000]}...")
        try:
            response = await loop.run_in_executor(
                None,
                lambda: asyncio.run(llm_config._llm(full_prompt, max_tokens=max_tokens, temperature=0.2, stop=None))
            )
            logger.debug(f"[LLM DEBUG] Raw LLM sectioned output for {file_path} (attempt {attempt}): {response[:200]}...")
        except Exception as e:
            logger.error(f"[LLM ERROR] Exception during LLM sectioned requirements for {file_path}: {e}")
            response = ""
        sections = _parse_sections(response)
        prev_output = response
        missing_sections = {req for req in REQUIRED_SECTIONS if req not in sections or not sections[req] or len(sections[req].strip()) < 10}
        logger.debug(f"Attempt {attempt} for {file_path}: missing sections: {missing_sections}")
        # If the LLM returned only a summary or rules, re-prompt with a warning
        if len(sections) == 1 and 'overview' in sections:
            logger.warning(f"LLM returned only a summary for {file_path}, re-prompting for all sections.")
            continue
        if not missing_sections:
            return {req: sections[req].strip() for req in REQUIRED_SECTIONS}
    logger.warning(f"LLM failed to generate all required sections for {file_path} after 3 attempts. Missing: {missing_sections}")
    return {"error": f"LLM failed to generate all required sections for {file_path}. Please check the file content or try again."}

# Utility functions (unchanged)
@lru_cache(maxsize=1000)
def count_tokens(text: str, model: str = "gpt-4o") -> int:
    try:
        encoding = tiktoken.encoding_for_model(model)
        tokens = encoding.encode(text)
        return len(tokens)
    except Exception:
        with log_lock:
            logger.warning(f"Token counting error for model {model}. Using fallback.")
        return int(len(text) / 4)

def retry_on_failure(max_attempts=2, delay=0.5):
    def decorator(func):
        @wraps(func)
        async def wrapper(*args, **kwargs):
            for attempt in range(max_attempts):
                try:
                    return await func(*args, **kwargs)
                except Exception as e:
                    with log_lock:
                        logger.error(f"Attempt {attempt + 1}/{max_attempts} failed: {str(e)}")
                    if attempt < max_attempts - 1:
                        await asyncio.sleep(delay)
                    else:
                        raise e
        return wrapper
    return decorator

async def clone_repository(repo_url: str, target_dir: str) -> bool:
    with log_lock:
        logger.info(f"Starting repository cloning: {repo_url} to {target_dir}")
    try:
        steps = ["Initializing", "Fetching metadata", "Downloading files", "Checking out branch"]
        for i, step in enumerate(steps, 1):
            with log_lock:
                logger.debug(f"Cloning progress: {step} ({i}/{len(steps)})")
            await asyncio.sleep(0.1)
        return True
    except Exception as e:
        with log_lock:
            logger.error(f"Failed to clone repository: {str(e)}")
        return False

async def generate_graph_from_requirement(requirement: str, target_graph: str = '', file_content: str = None) -> GraphResponse:
    if not file_content:
        file_content = requirement

    if count_tokens(file_content) > MAX_SAFE_TOKENS:
        with log_lock:
            logger.warning(f"File content for {target_graph}. Token limit exceeded. Truncating.")
        file_content = file_content[:MAX_SAFE_TOKENS * 3 // 4]

    if target_graph.lower() == 'entity relationship diagram':
        prompt = (
            "Create a precise Entity-Relationship Diagram with strict syntax rules:\n"
            f"Code Context:\n```\n{file_content}\n```\n\n"
            "STRICT GUIDELINES:\n"
            "1. ALWAYS use graph TD or graph LR\n"
            "2. Use ONLY square brackets [Entity]\n"
            "3. Connect with --> or ---\n"
            "4. NO special characters in node names\n"
            "5. Keep diagram SIMPLE and VALID\n"
            "Example:\n"
            "```mermaid\n"
            "graph TD\n"
            "    A[Customer] --> B[Order]\n"
            "    B --> C[Product]\n"
            "```\n"
            "Return ONLY valid Mermaid code."
        )
    elif target_graph.lower() == 'requirement diagram':
        prompt = (
            "Create a precise Mermaid Requirement Diagram with strict syntax rules:\n"
            f"Code Context:\n```\n{file_content}\n```\n\n"
            "STRICT GUIDELINES:\n"
            "1. ALWAYS use graph TD or graph LR\n"
            "2. Use ONLY square brackets [Requirement]\n"
            "3. Connect with --> or ---\n"
            "4. NO special characters in node names\n"
            "5. Keep diagram SIMPLE and VALID\n"
            "Example:\n"
            "```mermaid\n"
            "graph TD\n"
            "    A[User Authentication] --> B[Access Control]\n"
            "    B --> C[Data Validation]\n"
            "```\n"
            "Return ONLY valid Mermaid code."
        )
    else:
        return GraphResponse(
            target_graph=target_graph,
            generated_code=f"Unsupported graph type: {target_graph}"
        )

    async with llm_semaphore:
        try:
            async with asyncio.timeout(30):
                with ThreadPoolExecutor(max_workers=1) as executor:
                    result = await asyncio.get_event_loop().run_in_executor(
                        executor, lambda: asyncio.run(llm_config._llm(prompt))
                    )
            generated_code = result.strip()
            if not generated_code or len(generated_code.split('\n')) < 2:
                return GraphResponse(
                    target_graph=target_graph,
                    generated_code=f"Error: Insufficient graph definition for {target_graph}"
                )
            pattern = r"```mermaid\s*([\s\S]*?)\s*```"
            match = re.search(pattern, generated_code)
            if match:
                generated_code = match.group(1).strip()
            lines = generated_code.split('\n')
            if not lines[0].startswith('graph'):
                return GraphResponse(
                    target_graph=target_graph,
                    generated_code=f"Error: Invalid graph type. Must start with 'graph TD' or 'graph LR'"
                )
            return GraphResponse(target_graph=target_graph, generated_code=generated_code)
        except asyncio.TimeoutError:
            with log_lock:
                logger.error(f"Graph generation timed out for {target_graph}")
            return GraphResponse(
                target_graph=target_graph,
                generated_code=f"Generation Error: Timeout"
            )
        except Exception as e:
            with log_lock:
                logger.error(f"Graph generation error for {target_graph}: {e}")
            return GraphResponse(
                target_graph=target_graph,
                generated_code=f"Generation Error: {str(e)}"
            )

async def split_into_chunks(content: str, language: str, file_path: str, max_tokens: int = MAX_SAFE_TOKENS) -> List[Tuple[str, int, int]]:
    chunks = []
    current_line = 0
    lines = content.splitlines()
    total_lines = len(lines)
    total_tokens = count_tokens(content)
    with log_lock:
        logger.info(f"[CHUNK] Starting chunking for {file_path}: {total_tokens} tokens, {total_lines} lines")

    summary = "Summary unavailable.\n"
    if total_tokens < MAX_SAFE_TOKENS:
        summary_prompt = (
            f"Summarize the purpose and structure of the following {language} content in 50-100 words:\n"
            f"```\n{content[:1500]}\n[...truncated...]\n```\n"
            "Provide a concise overview of the file's role, main components, and key functionalities."
        )
        try:
            async with asyncio.timeout(5):
                with ThreadPoolExecutor(max_workers=1) as executor:
                    summary_result = await asyncio.get_event_loop().run_in_executor(
                        executor, lambda: asyncio.run(llm_config._llm(summary_prompt))
                    )
                summary = summary_result.strip() + "\n"
                with log_lock:
                    logger.info(f"[CHUNK] Generated summary for {file_path}")
        except asyncio.TimeoutError:
            with log_lock:
                logger.warning(f"[CHUNK] Summary generation timed out for {file_path}")
        except Exception as exc:
            with log_lock:
                logger.warning(f"[CHUNK] Failed to generate summary for {file_path}: {exc}")

    target_tokens = max_tokens * 4 // 5
    tokens_per_line = total_tokens / total_lines if total_lines > 0 else 1
    target_lines = int(target_tokens / tokens_per_line) if tokens_per_line > 0 else MAX_LINES_PER_CHUNK

    while current_line < total_lines:
        remaining_lines = total_lines - current_line
        chunk_lines = lines[current_line:current_line + min(target_lines, remaining_lines)]
        chunk_content = '\n'.join(chunk_lines)
        token_count = count_tokens(chunk_content)

        while token_count > max_tokens and len(chunk_lines) > 1:
            target_lines = max(1, target_lines * 3 // 4)
            chunk_lines = lines[current_line:current_line + target_lines]
            chunk_content = '\n'.join(chunk_lines)
            token_count = count_tokens(chunk_content)

        chunk_end = current_line + len(chunk_lines)
        chunk_content = f"File Summary:\n{summary}\nChunk Content:\n{chunk_content}"
        if count_tokens(chunk_content) > max_tokens:
            with log_lock:
                logger.warning(f"[CHUNK] Chunk at lines {current_line}-{chunk_end} in {file_path} exceeds token limit. Truncating.")
            chunk_lines = chunk_lines[:len(chunk_lines) * 3 // 4]
            chunk_content = f"File Summary:\n{summary}\nChunk Content:\n{chr(10).join(chunk_lines)}"

        chunks.append((chunk_content, current_line, chunk_end))
        with log_lock:
            logger.info(f"[CHUNK] Chunk {len(chunks)} for {file_path}: {token_count} tokens, lines {current_line}-{chunk_end}")
        current_line = chunk_end

    if not chunks:
        # Always emit at least one chunk (even for empty files)
        chunks.append(("", 1, 1))
        with log_lock:
            logger.warning(f"[CHUNK] No chunks created for {file_path}. Emitting empty chunk.")
    with log_lock:
        logger.info(f"[CHUNK] Completed chunking for {file_path}: {len(chunks)} chunks")
    return chunks

async def detect_business_rules(content: str, language: str) -> List[Dict[str, str]]:
    rules_detected = []
    domains = ["financial", "ecommerce", "general"]
    if language.lower() in ["excel", "vba"]:
        domains.append("excel")

    # Regex-based detection
    for domain in domains:
        config = RULES_CONFIG[domain]
        for pattern in config["patterns"]:
            matches = re.finditer(pattern, content, re.MULTILINE | re.IGNORECASE)
            for match in matches:
                rule_text = match.group(0)
                rules_detected.append({
                    "rule": rule_text,
                    "description": config["description"],
                    "domain": domain,
                    "source": "regex"
                })
                with log_lock:
                    logger.debug(f"Detected {domain} rule via regex: {rule_text[:100]}...")

    # Semantic-based detection
    if len(rules_detected) < 3:
        semantic_prompt = (
            f"Analyze the following {language} content for organization-specific business rules or logic not captured by regex patterns:\n"
            f"```\n{content[:2000]}\n[...truncated...]\n```\n"
            "Identify up to 3 organization-specific rules (e.g., proprietary calculations, custom workflows, or conditional logic). "
            "For each rule, provide a high-level description without revealing sensitive details. "
            "Return a JSON list of objects with 'rule', 'description', and 'domain' fields. If no rules are found, return an empty list."
        )

        try:
            async with llm_semaphore:
                async with asyncio.timeout(15):
                    with ThreadPoolExecutor(max_workers=1) as executor:
                        result = await asyncio.get_event_loop().run_in_executor(
                            executor, lambda: asyncio.run(llm_config._llm(semantic_prompt))
                        )
            output_text = result.strip()

            with log_lock:
                logger.debug(f"Raw LLM output: {output_text[:500]}...")

            if not output_text:
                raise ValueError("Empty response from LLM")

            semantic_rules = extract_json_from_llm_output(output_text)
            if not isinstance(semantic_rules, list):
                raise ValueError("LLM response is not a list")

            for rule in semantic_rules:
                rule["source"] = "semantic"
                rules_detected.append(rule)
                with log_lock:
                    logger.debug(f"Detected {rule.get('domain', 'unknown')} rule via semantic: {rule.get('description', '')[:100]}...")

        except (json.JSONDecodeError, ValueError, asyncio.TimeoutError, Exception) as e:
            with log_lock:
                logger.warning(f"Semantic rule detection failed: {str(e)} | Raw output: {output_text[:200]}")

    # Final log summary
    if not rules_detected:
        with log_lock:
            logger.debug(f"No business rules detected in content for language {language}")
    else:
        with log_lock:
            logger.info(f"Detected {len(rules_detected)} business rules for language {language}")

    return rules_detected

@retry_on_failure(max_attempts=2)
async def generate_requirements_for_chunk(chunk_content: str, chunk_index: int, language: str) -> str:
    """Generate requirements for a code chunk with enhanced RAG"""
    try:
        # Calculate tokens and handle limits
        chunk_tokens = count_tokens(chunk_content)
        chunk_hash = hashlib.md5(chunk_content.encode()).hexdigest()
        
        # Get relevant context from vector store with filtering
        context_data = await vector_store.get_related_context(
            query=chunk_content,
            k=3,  # Reduced from 5 to save tokens
            filters={"language": language}
        )
        
        # Calculate total tokens including context
        context_tokens = count_tokens(context_data.get('context', ''))
        total_tokens = chunk_tokens + context_tokens
        
        # If total tokens exceed limit, truncate context
        if total_tokens > MAX_SAFE_TOKENS:
            with log_lock:
                logger.warning(f"Total tokens ({total_tokens}) exceed limit. Truncating context.")
            # Keep chunk content but reduce context
            context_data['context'] = context_data.get('context', '')[:MAX_SAFE_TOKENS // 4]
        
        # If chunk itself is too large, truncate it
        if chunk_tokens > MAX_SAFE_TOKENS * 3 // 4:
            with log_lock:
                logger.warning(f"Chunk {chunk_index + 1} exceeds token limit. Truncating.")
            chunk_content = chunk_content[:MAX_SAFE_TOKENS * 3 // 4]
        
        # Detect business rules
        rules_detected = await detect_business_rules(chunk_content, language)
        rules_summary = (
            "\n".join([f"Detected {rule['domain']} rule ({rule['source']}): {rule['description']}" for rule in rules_detected])
            if rules_detected else ""
        )
        
        # Build enhanced prompt with metadata
        prompt = f"""Act as a senior business analyst reviewing a {language} source file segment.
This is segment {chunk_index + 1} of a larger file. Generate concise business requirements in plain text, avoiding markdown symbols (*, -, #, **).
Do NOT use the phrase "this chunk" or "this segment".

The output MUST have five sections: Overview, Objective, Use Case, Key Functionalities, Workflow Summary, separated by two newlines.

If the code contains organization-specific logic or calculations (e.g., proprietary formulas or business rules), generalize them without revealing sensitive details. For example, instead of specifying an exact calculation, say "performs proprietary calculations based on internal rules."

Use the following detected rules for context:
{rules_summary}"""

        # Add context if available
        if context_data.get('context'):
            metadata = context_data.get('metadata', {})
            related_files = context_data.get('related_files', [])
            prompt += f"""

Relevant Context from Similar Code:
{context_data['context']}

Context Metadata:
- Total related files: {metadata.get('total_files', 0)}
- Most relevant file: {related_files[0]['file_path'] if related_files else 'None'}"""

        prompt += f"""

Analyze this segment:
```{language}
{chunk_content}
```

Constraints:
- Plain text, no markdown
- Five non-empty sections
- Key Functionalities numbered, each followed by a blank line
- Use file summary and detected rules for context
- Return only the structured requirements

Generate the following sections:
Overview
Describe the segment's role in the file's purpose and system context (50-100 words).

Objective
State the business goal or problem the segment addresses (1-2 sentences).

Use Case
Outline business scenarios where the segment's functionality is critical (2-3 sentences).

Key Functionalities
List 2-5 core business capabilities, each on a new line, numbered (e.g., 1., 2.), followed by a blank line. Each functionality must include a brief description.

Workflow Summary
Describe how the segment supports the file's business process (2-3 sentences).
"""
        
        async with llm_semaphore:
            try:
                async with asyncio.timeout(30):
                    with ThreadPoolExecutor(max_workers=1) as executor:
                        result = await asyncio.get_event_loop().run_in_executor(
                            executor, lambda: asyncio.run(llm_config._llm(prompt, max_tokens=500, temperature=0.2))
                        )
                
                req_text = result.strip()
                with log_lock:
                    logger.debug(f"Raw LLM output for chunk {chunk_index + 1} (hash: {chunk_hash}): {req_text[:200]}...")
                
                # Process and format the response
                req_text = re.sub(r'\n\s*\n+', '\n\n', req_text)
                req_text = re.sub(r'this chunk|this segment', 'the code', req_text, flags=re.IGNORECASE)
                
                # Split into sections
                sections = re.split(r'\n\n(?=Overview|Objective|Use Case|Key Functionalities|Workflow Summary)', req_text)
                section_dict = {}
                
                for section in sections:
                    for header in ['Overview', 'Objective', 'Use Case', 'Key Functionalities', 'Workflow Summary']:
                        if section.startswith(header):
                            section_dict[header] = section[len(header):].strip()
                            break
                
                # Ensure all sections exist
                for header in ['Overview', 'Objective', 'Use Case', 'Key Functionalities', 'Workflow Summary']:
                    if header not in section_dict or not section_dict[header]:
                        with log_lock:
                            logger.warning(f"Missing or empty {header} in chunk {chunk_index + 1} (hash: {chunk_hash})")
                        section_dict[header] = (
                            f"Supports {language.lower()} file operations." if header == 'Overview' else
                            f"To enable core {language.lower()} file functionality." if header == 'Objective' else
                            f"Supports general business tasks in {language.lower()} applications." if header == 'Use Case' else
                            "1. Basic Processing: Handles core tasks.\n\n2. Support Functions: Assists operations." if header == 'Key Functionalities' else
                            f"Contributes to the {language.lower()} file's workflow."
                        )
                
                # Format Key Functionalities
                if section_dict['Key Functionalities']:
                    func_lines = section_dict['Key Functionalities'].split('\n')
                    formatted_funcs = []
                    for line in func_lines:
                        if re.match(r'^\d+\.\s', line):
                            formatted_funcs.append(line.strip())
                    if not formatted_funcs:
                        formatted_funcs = ["1. Basic Processing: Handles core tasks.", "2. Support Functions: Assists operations."]
                    section_dict['Key Functionalities'] = '\n\n'.join(formatted_funcs)
                
                # Combine sections
                output_sections = [
                    f"Overview\n{section_dict['Overview']}",
                    f"Objective\n{section_dict['Objective']}",
                    f"Use Case\n{section_dict['Use Case']}",
                    f"Key Functionalities\n{section_dict['Key Functionalities']}",
                    f"Workflow Summary\n{section_dict['Workflow Summary']}"
                ]
                
                return '\n\n'.join(output_sections)
                
            except asyncio.TimeoutError:
                with log_lock:
                    logger.error(f"LLM timed out for chunk {chunk_index + 1} (hash: {chunk_hash})")
                return (
                    f"Overview\nAnalysis limited due to timeout for {language.lower()} file operations.\n\n"
                    f"Objective\nTo support core {language.lower()} file functionality.\n\n"
                    f"Use Case\nSupports general {language.lower()} business tasks.\n\n"
                    f"Key Functionalities\n1. Partial Processing: Supports basic operations.\n\n2. Fallback Support: Provides minimal functionality.\n\n"
                    f"Workflow Summary\nContributes to {language.lower()} file workflow."
                )
            except Exception as e:
                with log_lock:
                    logger.error(f"Error for chunk {chunk_index + 1}: {e}")
                raise e
    except Exception as e:
        with log_lock:
            logger.error(f"Error in generate_requirements_for_chunk: {str(e)}")
        return (
            f"Overview\nError analyzing {language.lower()} code segment.\n\n"
            f"Objective\nUnable to determine specific objectives.\n\n"
            f"Use Case\nGeneral {language.lower()} file operations.\n\n"
            f"Key Functionalities\n1. Basic Processing: Handles core tasks.\n\n2. Error Recovery: Provides fallback functionality.\n\n"
            f"Workflow Summary\nSupports basic {language.lower()} file workflow."
        )

@retry_on_failure(max_attempts=2)
async def extract_organization_specific_rules_for_chunk(chunk_content: str, chunk_index: int, language: str) -> str:
    chunk_hash = hashlib.md5(chunk_content.encode()).hexdigest()
    if count_tokens(chunk_content) > MAX_SAFE_TOKENS:
        with log_lock:
            logger.warning(f"Chunk {chunk_index + 1} exceeds token limit for org-specific rules. Truncating.")
        chunk_content = chunk_content[:MAX_SAFE_TOKENS * 3 // 4]

    rules_detected = await detect_business_rules(chunk_content, language)
    rules_summary = (
        "\n".join([f"Detected {rule['domain']} rule ({rule['source']}): {rule['description']}" for rule in rules_detected])
        if rules_detected else ""
    )

    min_lines = 3
    if len(chunk_content) > 1000:
        min_lines = 5
    elif len(chunk_content) > 5000:
        min_lines = 7
    elif len(chunk_content) > 10000:
        min_lines = 10

    prompt = (
        f"You are a senior business analyst. Carefully review the following {language} source file segment and produce a highly detailed, elaborate, and comprehensive business logic summary.\n"
        f"For each section below (Overview, Objective, Use Case, Key Functionalities, Workflow Summary), provide an exhaustive, content-rich explanation based strictly on the actual file content.\n"
        f"- Expand on every available detail: describe all functions, classes, workflows, and business scenarios in depth.\n"
        f"- Avoid vague, generic, or filler statements. Do NOT repeat phrases like 'utility file' or 'supports operations' unless those are explicitly and uniquely justified by the file content.\n"
        f"- For small files, elaborate as much as possible on every element, inferring business purpose, logic, and workflow from names, comments, and code structure.\n"
        f"- For each rule, provide a high-level description of its purpose and usage (e.g., 'Calculates loan eligibility using a proprietary formula based on customer data').\n"
        f"- Use the following detected rules for context:\n{rules_summary}\n"
        f"- Your output MUST be at least {min_lines} lines, with every line relevant and derived from the file. If the file is small, expand each section as much as possible based on available information.\n"
        f"- Do NOT add generic, filler, or placeholder lines. If there is insufficient business logic, only output what is truly present in the file, but do your best to elaborate on every detail.\n"
        f"If no organization-specific rules are found, return \"\"\n"
        f"Return only the summarized rules.\n"
        f"Analyze this segment:\n"
        f"```{language}\n{chunk_content}\n```\n"
    )
    async with llm_semaphore:
        try:
            async with asyncio.timeout(30):
                with ThreadPoolExecutor(max_workers=1) as executor:
                    result = await asyncio.get_event_loop().run_in_executor(
                        executor, lambda: asyncio.run(llm_config._llm(prompt))
                    )
            rules_text = result.strip()
            with log_lock:
                logger.debug(f"Raw LLM output for org-specific rules chunk {chunk_index + 1} (hash: {chunk_hash}): {rules_text[:200]}...")

            rules_text = re.sub(r'\n\s*\n+', '\n\n', rules_text)
            rules_text = re.sub(r'this chunk|this segment', 'the content', rules_text, flags=re.IGNORECASE)
            if not rules_text or rules_text.isspace():
                rules_text = ""
            return rules_text
        except asyncio.TimeoutError:
            with log_lock:
                logger.error(f"LLM timed out for org-specific rules chunk {chunk_index + 1} (hash: {chunk_hash})")
            return ""
        except Exception as e:
            with log_lock:
                logger.error(f"Error for org-specific rules chunk {chunk_index + 1}: {e}")
            return ""

async def combine_requirements(requirements_list: List[str], language: str) -> str:
    if not requirements_list:
        with log_lock:
            logger.error("No requirements provided to combine")
        return (
            f"Overview\nNo analysis available for {language.lower()} file.\n\n"
            f"Objective\nTo support intended {language.lower()} business functions.\n\n"
            f"Use Case\nIntended {language.lower()} business operations.\n\n"
            f"Key Functionalities\n1. Intended Functionality: Expected to provide core features.\n\n2. Support Functions: Assists operations.\n\n"
            f"Workflow Summary\nExpected to integrate with {language.lower()} system workflows."
        )

    overviews = []
    objectives = []
    use_cases = []
    functionalities = []
    workflows = []
    func_number = 1
    seen_funcs = set()

    with log_lock:
        logger.debug(f"Combining {len(requirements_list)} chunk requirements for {language} file")

    for req_text in requirements_list:
        try:
            # Updated regex to split on Markdown or plain section headers
            sections = re.split(r'\n+\s*#*\s*(?=Overview|Objective|Use Case|Key Functionalities|Workflow Summary)', req_text, flags=re.IGNORECASE)
            section_dict = {}
            for section in sections:
                for header in ['Overview', 'Objective', 'Use Case', 'Key Functionalities', 'Workflow Summary']:
                    if _normalize_header(section).startswith(header.lower()):
                        section_dict[header] = section[len(header):].strip()
                        break

            if len(section_dict) != 5:
                with log_lock:
                    logger.warning(f"Invalid chunk requirements format: {req_text[:200]}...")
                continue

            overview = section_dict['Overview']
            objective = section_dict['Objective']
            use_case = section_dict['Use Case']
            func_text = section_dict['Key Functionalities']
            workflow = section_dict['Workflow Summary']

            if overview:
                overviews.append(overview)
            if objective:
                objectives.append(objective)
            if use_case:
                use_cases.append(use_case)
            if workflow:
                workflows.append(workflow)

            func_lines = func_text.split('\n\n')
            for func in func_lines:
                if re.match(r'^\d+\.\s', func):
                    parts = func.split(':', 1)
                    if len(parts) == 2 and parts[1].strip() not in seen_funcs:
                        seen_funcs.add(parts[1].strip())
                        functionalities.append(f"{func_number}. {parts[1].strip()}")
                        func_number += 1
        except Exception as e:
            with log_lock:
                logger.warning(f"Error parsing chunk requirements: {str(e)}")

    def summarize_section(items, max_tokens, section_name):
        combined = " ".join(set(item for item in items if item))
        sentences = re.split(r'(?<=[.!?])\s+', combined)
        result = []
        current_tokens = 0
        for sentence in sentences:
            sentence_tokens = count_tokens(sentence)
            if current_tokens + sentence_tokens <= max_tokens and sentence.strip():
                result.append(sentence.strip())
                current_tokens += sentence_tokens
            if current_tokens >= max_tokens:
                break
        summary = " ".join(result)
        if not summary or count_tokens(summary) > max_tokens:
            with log_lock:
                logger.warning(f"Summary for {section_name} exceeded token limit or empty. Using fallback.")
            return (
                f"Supports {language.lower()} file operations." if section_name == 'Overview' else
                f"To enable core {language.lower()} business functionality." if section_name == 'Objective' else
                f"Supports {language.lower()} business scenarios." if section_name == 'Use Case' else
                f"Integrates with {language.lower()} file workflows."
            )
        return summary

    overview = summarize_section(overviews, MAX_SECTION_TOKENS, 'Overview')
    objective = summarize_section(objectives, MAX_SECTION_TOKENS // 2, 'Objective')
    use_case = summarize_section(use_cases, MAX_SECTION_TOKENS // 2, 'Use Case')
    workflow = summarize_section(workflows, MAX_SECTION_TOKENS // 2, 'Workflow Summary')

    max_funcs = 10
    functionalities = functionalities[:max_funcs]
    functionalities_section = ''
    if functionalities:
        functionalities_section = f"Key Functionalities\n{'\n\n'.join(functionalities)}\n\n"
    # else: leave functionalities_section empty (or optionally: functionalities_section = 'Key Functionalities\nNo explicit functionalities detected.\n\n')

    combined = (
        f"Overview\n{overview}\n\n"
        f"Objective\n{objective}\n\n"
        f"Use Case\n{use_case}\n\n"
        f"{functionalities_section}"
        f"Workflow Summary\n{workflow}"
    )

    total_tokens = count_tokens(combined)
    if total_tokens > MAX_SAFE_TOKENS:
        with log_lock:
            logger.warning(f"Combined requirements exceed token limit ({total_tokens}). Summarizing.")
        summary_prompt = (
            f"Summarize the following {language} requirements into a concise document under {MAX_SAFE_TOKENS} tokens:\n"
            f"```\n{combined}\n```\n"
            "Maintain structure with Overview, Objective, Use Case, Key Functionalities, Workflow Summary, separated by two newlines. "
            "Ensure complete sentences, remove redundancy, use clear language, avoid markdown symbols. "
            "Limit Key Functionalities to 5 numbered items (e.g., 1., 2.), each followed by a blank line."
        )
        try:
            async with asyncio.timeout(10):
                with ThreadPoolExecutor(max_workers=1) as executor:
                    result = await asyncio.get_event_loop().run_in_executor(
                        executor, lambda: asyncio.run(llm_config._llm(summary_prompt))
                    )
                combined = result.strip()
                sections = re.split(r'\n\n(?=Overview|Objective|Use Case|Key Functionalities|Workflow Summary)', combined)
                section_dict = {}
                for section in sections:
                    for header in ['Overview', 'Objective', 'Use Case', 'Key Functionalities', 'Workflow Summary']:
                        if section.startswith(header):
                            section_dict[header] = section[len(header):].strip()
                            break
                for header in ['Overview', 'Objective', 'Use Case', 'Workflow Summary']:
                    if header in section_dict and section_dict[header]:
                        sentences = re.split(r'(?<=[.!?])\s+', section_dict[header])
                        section_dict[header] = " ".join(sentences[:-1]) if sentences and not sentences[-1].endswith(('.', '!', '?')) else section_dict[header]
                if 'Key Functionalities' in section_dict:
                    func_lines = section_dict['Key Functionalities'].split('\n')
                    formatted_funcs = [line.strip() for line in func_lines if re.match(r'^\d+\.\s', line)][:5]
                    section_dict['Key Functionalities'] = '\n\n'.join(formatted_funcs) if formatted_funcs else '\n\n'.join(functionalities[:5])
                combined = (
                    f"Overview\n{section_dict.get('Overview', f'Supports {language.lower()} file operations.')}\n\n"
                    f"Objective\n{section_dict.get('Objective', f'To enable core {language.lower()} business functionality.')}\n\n"
                    f"Use Case\n{section_dict.get('Use Case', f'Supports {language.lower()} business scenarios.')}\n\n"
                    f"Key Functionalities\n{section_dict.get('Key Functionalities', '\n\n'.join(functionalities[:5]))}\n\n"
                    f"Workflow Summary\n{section_dict.get('Workflow Summary', f'Integrates with {language.lower()} file workflows.')}"
                )
        except asyncio.TimeoutError:
            with log_lock:
                logger.error("Summary timed out. Truncating to complete sentences.")
            combined = (
                f"Overview\n{summarize_section(overviews, MAX_SECTION_TOKENS // 2, 'Overview')}\n\n"
                f"Objective\n{summarize_section(objectives, MAX_SECTION_TOKENS // 4, 'Objective')}\n\n"
                f"Use Case\n{summarize_section(use_cases, MAX_SECTION_TOKENS // 4, 'Use Case')}\n\n"
                f"Key Functionalities\n{'\n\n'.join(functionalities[:5])}\n\n"
                f"Workflow Summary\n{summarize_section(workflows, MAX_SECTION_TOKENS // 4, 'Workflow Summary')}"
            )

    with log_lock:
        logger.debug(f"Combined requirements: {count_tokens(combined)} tokens")
    return combined

@lru_cache(maxsize=100)
def _hash_content(content: str) -> str:
    return hashlib.md5(content.encode()).hexdigest()

async def extract_vba_macros(file_path: str) -> str:
    """Extract VBA macros from .xlsm, .xlsb, or .xls files using oletools or fallback methods. Also parse macro names, parameters, and comments for richer summaries."""
    try:
        import subprocess
        import sys
        import os
        import re
        
        # Check if oletools is installed
        try:
            import oletools.olevba as olevba
            vba_parser = olevba.VBA_Parser(file_path)
            vba_modules = []
            
            if vba_parser.detect_vba_macros():
                for (filename, stream_path, vba_filename, vba_code) in vba_parser.extract_macros():
                    if vba_code.strip():
                        # Parse macro names, parameters, and comments
                        macro_summaries = []
                        for match in re.finditer(r'((Sub|Function)\s+(\w+)\s*\((.*?)\))', vba_code, re.IGNORECASE):
                            header, kind, name, params = match.groups()
                            # Extract comments before macro
                            pre_lines = vba_code[:match.start()].splitlines()[-3:]
                            comments = [l.strip() for l in pre_lines if l.strip().startswith("'")]
                            macro_summaries.append(f"Macro: {name}\nType: {kind}\nParameters: {params}\nComments: {' '.join(comments) if comments else '[No comments]'}\n")
                        vba_modules.append(f"--- VBA Module: {vba_filename} ---\n{vba_code}\n\nSummary:\n" + "\n".join(macro_summaries))
                vba_parser.close()
                
                if vba_modules:
                    return "\n\n".join(vba_modules)
                return "No VBA macros found or could not be extracted."
            else:
                return "No VBA macros detected in the file."
        except ImportError:
            # Fallback to basic detection without oletools
            if file_path.lower().endswith(('.xlsm', '.xlsb', '.xls')):
                try:
                    # Attempt to read VBA project data directly (basic approach)
                    with open(file_path, 'rb') as f:
                        content = f.read().decode('latin1', errors='ignore')
                        if 'Sub ' in content or 'Function ' in content:
                            return "VBA macros detected (detailed analysis requires oletools package: pip install oletools)"
                        return "No VBA macros detected in the file."
                except Exception as e:
                    with log_lock:
                        logger.warning(f"Fallback VBA detection failed for {file_path}: {str(e)}")
                    return "Error extracting VBA macros: oletools not installed and fallback detection failed."
        except Exception as e:
            with log_lock:
                logger.warning(f"Error extracting VBA macros from {file_path}: {str(e)}")
            return f"Error extracting VBA macros: {str(e)}"
            
    except Exception as e:
        with log_lock:
            logger.error(f"Unexpected error in VBA macro extraction for {file_path}: {str(e)}")
        return f"Error extracting VBA macros: {str(e)}"

async def extract_excel_content(file_path: str) -> str:
    """Extract content from Excel files (.xlsx, .xlsm, .xls) including VBA macros, formulas, validations, and formatting."""
    try:
        content_lines = []
        _, ext = os.path.splitext(file_path)
        is_macro_enabled = ext.lower() in {'.xlsm', '.xlsb', '.xls'}

        # Always try to extract VBA macros for macro-enabled files (even if workbook fails)
        vba_content = None
        if is_macro_enabled:
            try:
                vba_content = await extract_vba_macros(file_path)
                if vba_content and not vba_content.startswith("Error"):
                    content_lines.append("=== VBA MACROS ===")
                    content_lines.append(vba_content)
                    content_lines.append("")
                elif vba_content and vba_content.startswith("Error"):
                    content_lines.append(f"Error extracting VBA macros: {vba_content}")
                    content_lines.append("")
            except Exception as e:
                content_lines.append(f"Error extracting VBA macros: {str(e)}")
                content_lines.append("")

        # Try to open the workbook for reading
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
        except Exception as e:
            content_lines.append(f"Error reading workbook content: {str(e)}")
            # If we got any VBA, at least return that
            if content_lines:
                return "\n".join(content_lines)
            else:
                return f"Error reading workbook content: {str(e)}"

        
        try:
            content_lines.append("=== WORKSHEET DATA ===")
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                content_lines.append(f"\n--- Sheet: {sheet_name} ---")
                
                # Extract cell values and formulas (limit to 100 rows to avoid performance issues)
                max_rows = min(100, sheet.max_row) if sheet.max_row else 100
                for row in sheet.iter_rows(max_row=max_rows):
                    for cell in row:
                        if cell.value is not None:
                            cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
                            if cell.data_type == 'f':  # Formula
                                content_lines.append(f"FORMULA: {cell_ref} = {cell.value}")
                            else:
                                content_lines.append(f"CELL: {cell_ref} = {cell.value}")
                
                # Extract data validation rules
                if hasattr(sheet, 'data_validations') and sheet.data_validations.dataValidation:
                    content_lines.append("\n  Data Validations:")
                    for dv in sheet.data_validations.dataValidation:
                        try:
                            range_str = dv.sqref.ranges[0].coord if hasattr(dv, 'sqref') and dv.sqref else 'Unknown range'
                            validation_type = getattr(dv, 'type', 'custom')
                            formula1 = getattr(dv, 'formula1', None)
                            formula2 = getattr(dv, 'formula2', None)
                            
                            validation_info = [f"Range: {range_str}", f"Type: {validation_type}"]
                            if formula1:
                                validation_info.append(f"Formula1: {formula1}")
                            if formula2:
                                validation_info.append(f"Formula2: {formula2}")
                                
                            content_lines.append(f"DATA_VALIDATION: {' | '.join(validation_info)}")
                        except Exception as e:
                            with log_lock:
                                logger.warning(f"Error processing data validation in {sheet_name}: {str(e)}")
                
                # Extract conditional formatting
                if hasattr(sheet, 'conditional_formatting') and sheet.conditional_formatting:
                    content_lines.append("\n  Conditional Formatting:")
                    for range_str, rules in sheet.conditional_formatting._cf_rules.items():
                        for rule in rules:
                            try:
                                rule_type = rule.type
                                formula = getattr(rule, 'formula', None)
                                if formula:
                                    if isinstance(formula, (list, tuple)):
                                        formula = "; ".join([str(f) for f in formula if f])
                                    content_lines.append(
                                        f"CONDITIONAL_FORMATTING: Range: {range_str} | Type: {rule_type} | Formula: {formula}"
                                    )
                            except Exception as e:
                                with log_lock:
                                    logger.warning(f"Error processing conditional formatting in {sheet_name}: {str(e)}")
                
                # Extract named ranges
                if hasattr(workbook, 'defined_names') and workbook.defined_names:
                    content_lines.append("\n  Named Ranges:")
                    if hasattr(workbook.defined_names, 'definedName'):  # Old versions
                        for name in workbook.defined_names.definedName:
                            content_lines.append(f"  - {name.name}: {name.value}")
                    else:  # New versions (4.0+)
                        for name, named_range in workbook.defined_names.items():
                            content_lines.append(f"  - {name}: {named_range.value}")
            
            # Extract pivot tables
            for sheet in workbook.worksheets:
                if hasattr(sheet, '_pivots') and sheet._pivots:
                    content_lines.append(f"\nPivot Tables in {sheet.title}:")
                    for pivot in sheet._pivots:
                        content_lines.append(f"  - Name: {pivot.name}")
                        content_lines.append(f"    Data Source: {pivot.cache.cacheSource.worksheetSource.ref}")
            
            # Extract external links
            if hasattr(workbook, 'external_links') and workbook.external_links:
                content_lines.append("\nExternal Links:")
                for link in workbook.external_links:
                    content_lines.append(f"  - {link.target}")
            
            # Extract data connections
            if hasattr(workbook, 'connections') and workbook.connections:
                content_lines.append("\nData Connections:")
                for conn in workbook.connections:
                    content_lines.append(f"  - {conn.name}: {conn.connection_string}")
            
        finally:
            workbook.close()
            
        return "\n".join(content_lines) if content_lines else "No content extracted from Excel file."
        
    except Exception as e:
        with log_lock:
            logger.error(f"Error extracting Excel content from {file_path}: {str(e)}")
        return f"Error extracting Excel content: {str(e)}"

async def extract_organization_specific_rules(file_path: str, base_dir: str, language: str) -> OrgSpecificRules:
    try:
        with log_lock:
            logger.debug(f"Extracting org-specific rules for file: {file_path}, Language: {language}")

        _, ext = os.path.splitext(file_path)
        is_excel = ext.lower() in {'.xlsx', '.xlsm', '.xls', '.xlsb'}
        
        if is_excel:
            language = "Excel" if ext.lower() == '.xlsx' else "VBA"
            content = await extract_excel_content(file_path)
        else:
            code_extensions = {'.py', '.js', '.jsx', '.ts', '.tsx', '.java', '.scala', '.rb', '.go', '.cpp', '.c', '.rs', '.kt', '.swift', '.php', '.cs'}
            if ext.lower() not in code_extensions:
                with log_lock:
                    logger.info(f"Skipping non-code file for org-specific rules: {file_path} (Extension: {ext})")
                return OrgSpecificRules(
                    relative_path=os.path.relpath(file_path, base_dir),
                    file_name=os.path.basename(file_path),
                    rules="Non-code file: Organization-specific rules extraction not applicable."
                )
            language = language or "Plain Text"
            async with aiofiles.open(file_path, mode='r', encoding='latin-1') as file:
                content = await file.read()

        total_tokens = count_tokens(content)
        with log_lock:
            logger.debug(f"File: {file_path}, Language: {language}, Size: {len(content)} bytes, Tokens: {total_tokens}")

        if total_tokens < 1000:
            content_hash = _hash_content(content)
            rules_text = await extract_organization_specific_rules_for_chunk(content, 0, language)
            with log_lock:
                logger.debug(f"Extracted org-specific rules for small file {file_path} (hash: {content_hash})")
            return OrgSpecificRules(
                relative_path=os.path.relpath(file_path, base_dir),
                file_name=os.path.basename(file_path),
                rules=rules_text
            )

        if total_tokens <= MAX_SAFE_TOKENS:
            if total_tokens > MAX_SAFE_TOKENS // 2:
                content = content[:int(len(content) * MAX_SAFE_TOKENS * 0.3 / total_tokens)]
                with log_lock:
                    logger.warning(f"Truncated {file_path} to {len(content)} bytes for org-specific rules")
            rules_text = await extract_organization_specific_rules_for_chunk(content, 0, language)
            return OrgSpecificRules(
                relative_path=os.path.relpath(file_path, base_dir),
                file_name=os.path.basename(file_path),
                rules=rules_text
            )
        else:
            with log_lock:
                logger.debug(f"Splitting {file_path} into chunks for org-specific rules due to high token count: {total_tokens}")
            try:
                chunks = await split_into_chunks(content, language, file_path)
                with log_lock:
                    logger.debug(f"Generated {len(chunks)} chunks for {file_path}")
                rules_list = []
                for i, (chunk_content, start_line, end_line) in enumerate(chunks):
                    try:
                        rules = await extract_organization_specific_rules_for_chunk(chunk_content, i, language)
                        if rules.strip():
                            rules_list.append(f"Chunk {i+1} (Lines {start_line}-{end_line}):\n{rules}")
                        if (i + 1) % 5 == 0 or i + 1 == len(chunks):
                            with log_lock:
                                logger.info(f"Processed chunk {i+1}/{len(chunks)} for org-specific rules in {file_path}")
                    except Exception as e:
                        with log_lock:
                            logger.error(f"Error processing chunk {i+1} for org-specific rules in {file_path}: {e}")
                        rules_list.append(f"Chunk {i+1} (Lines {start_line}-{end_line}):\nError extracting rules: {str(e)}")
                rules_text = "\n\n".join(rules_list) if rules_list else ""
            except MemoryError as me:
                with log_lock:
                    logger.error(f"Memory error during chunking for org-specific rules in {file_path}: {me}")
                rules_text = "Error extracting organization-specific rules due to memory constraints."
            except Exception as e:
                with log_lock:
                    logger.error(f"Chunking failed for org-specific rules in {file_path}: {str(e)}")
                rules_text = f"Error extracting organization-specific rules: {str(e)}"

        return OrgSpecificRules(
            relative_path=os.path.relpath(file_path, base_dir),
            file_name=os.path.basename(file_path),
            rules=rules_text
        )
    except Exception as e:
        with log_lock:
            logger.error(f"Error extracting org-specific rules for {file_path}: {str(e)}")
        return OrgSpecificRules(
            relative_path=os.path.relpath(file_path, base_dir),
            file_name=os.path.basename(file_path),
            rules=f"Error extracting organization-specific rules: {str(e)}"
        )

async def generate_project_summary(files_requirements: FilesRequirements) -> Tuple[str, List[GraphResponse]]:
    try:
        with log_lock:
            logger.info("Starting comprehensive project summary generation")
        
        workflow_summaries = []
        file_contexts = []
        
        for file_req in files_requirements.files:
            if file_req.requirements:
                sections = re.split(
                    r'\n\n(?=Overview|Objective|Use Case|Key Functionalities|Workflow Summary|Dependent Files|Technical Requirements|Business Logic)', 
                    file_req.requirements
                )
                
                file_info = {
                    'name': file_req.file_name,
                    'workflow': '',
                    'overview': '',
                    'objective': '',
                    'key_functions': '',
                    'use_case': ''
                }
                
                for section in sections:
                    section_lower = section.lower()
                    if section.startswith("Workflow Summary"):
                        workflow_text = section[len("Workflow Summary"):].strip()
                        if workflow_text and not workflow_text.startswith("Error"):
                            file_info['workflow'] = workflow_text
                            workflow_summaries.append(f"{file_req.file_name}: {workflow_text}")
                    elif section.startswith("Overview"):
                        file_info['overview'] = section[len("Overview"):].strip()
                    elif section.startswith("Objective"):
                        file_info['objective'] = section[len("Objective"):].strip()
                    elif section.startswith("Key Functionalities"):
                        file_info['key_functions'] = section[len("Key Functionalities"):].strip()
                    elif section.startswith("Use Case"):
                        file_info['use_case'] = section[len("Use Case"):].strip()
                
                if any(file_info.values()):
                    file_contexts.append(file_info)
        
        if not workflow_summaries and not file_contexts:
            with log_lock:
                logger.warning("No valid content found for project summary")
            return "Project analysis completed with limited information available for comprehensive summary generation.", []
        
        combined_context = ""
        
        if workflow_summaries:
            combined_context += "WORKFLOW SUMMARIES:\n" + "\n".join(workflow_summaries) + "\n\n"
        
        if file_contexts:
            combined_context += "ADDITIONAL FILE CONTEXT:\n"
            for file_info in file_contexts:
                if file_info['overview'] or file_info['objective'] or file_info['key_functions']:
                    combined_context += f"\n{file_info['name']}:\n"
                    if file_info['overview']:
                        combined_context += f"  Overview: {file_info['overview']}\n"
                    if file_info['objective']:
                        combined_context += f"  Objective: {file_info['objective']}\n"
                    if file_info['key_functions']:
                        combined_context += f"  Key Functions: {file_info['key_functions']}\n"
                    if file_info['use_case']:
                        combined_context += f"  Use Case: {file_info['use_case']}\n"
        
        if count_tokens(combined_context) > MAX_SAFE_TOKENS:
            with log_lock:
                logger.warning("Combined context exceeds token limit. Truncating for project summary.")
            combined_context = combined_context[:MAX_SAFE_TOKENS * 3 // 4]
        
        prompt = f"""
Act as a senior technical architect and business analyst conducting a comprehensive project review.
Based on the provided workflow summaries and file analysis, generate a detailed, well-structured project summary that follows this exact format:
## PROJECT OVERVIEW
[2-3 sentences describing the project's main purpose, domain, and scope]
## BUSINESS CONTEXT & OBJECTIVES
[2-3 sentences explaining the business problem being solved and strategic objectives]
## SYSTEM ARCHITECTURE & APPROACH
[3-4 sentences describing the technical approach, key architectural patterns, and system design principles]
## KEY CAPABILITIES & FEATURES
[4-5 sentences detailing the main functionalities, core features, and what the system can accomplish]
## WORKFLOW & PROCESS INTEGRATION
[3-4 sentences explaining how different components work together, data flow, and process coordination]
## BUSINESS VALUE & IMPACT
[2-3 sentences highlighting the expected business benefits, efficiency gains, and organizational impact]
## TECHNICAL FOUNDATION
[2-3 sentences covering the technology stack, integration patterns, and scalability considerations]
Guidelines:
- Write in clear, professional language suitable for both technical and business stakeholders
- Focus on business value while maintaining technical accuracy
- Avoid mentioning specific file names or implementation details
- Each section should be substantive and informative
- Use present tense and active voice
- Ensure logical flow between sections
- Keep the tone analytical and objective
Project Information:
{combined_context}
"""
        
        async with llm_semaphore:
            try:
                async with asyncio.timeout(45):
                    with ThreadPoolExecutor(max_workers=1) as executor:
                        result = await asyncio.get_event_loop().run_in_executor(
                            executor, lambda: asyncio.run(llm_config._llm(prompt))
                        )
                
                project_summary = result.strip()
                
                required_sections = [
                    "PROJECT OVERVIEW",
                    "BUSINESS CONTEXT & OBJECTIVES", 
                    "SYSTEM ARCHITECTURE & APPROACH",
                    "KEY CAPABILITIES & FEATURES",
                    "WORKFLOW & PROCESS INTEGRATION",
                    "BUSINESS VALUE & IMPACT",
                    "TECHNICAL FOUNDATION"
                ]
                
                missing_sections = [section for section in required_sections if section not in project_summary]
                
                if missing_sections or len(project_summary.split()) < 150:
                    with log_lock:
                        logger.warning(f"Generated summary may be incomplete. Missing sections: {missing_sections}")
                    
                    project_summary = f"""
## PROJECT OVERVIEW
This project implements a comprehensive software solution designed to address complex business requirements through integrated system components. The solution encompasses multiple modules working in coordination to deliver automated business processes and data management capabilities.
## BUSINESS CONTEXT & OBJECTIVES
The system addresses organizational needs for streamlined operations and improved process efficiency. It aims to reduce manual intervention while providing reliable, scalable solutions for business-critical operations.
## SYSTEM ARCHITECTURE & APPROACH
The architecture follows modular design principles with clear separation of concerns across different functional areas. Components are designed for maintainability and extensibility, utilizing established patterns for reliable system integration and data processing workflows.
## KEY CAPABILITIES & FEATURES
The system provides automated workflow processing, data validation and transformation, integrated reporting capabilities, and comprehensive error handling. It supports configurable business rules, multi-step process orchestration, and real-time monitoring of system operations.
## WORKFLOW & PROCESS INTEGRATION
Different system components communicate through well-defined interfaces, ensuring data consistency and process reliability. The workflow engine coordinates activities across modules, managing dependencies and ensuring proper sequencing of operations.
## BUSINESS VALUE & IMPACT
Implementation delivers improved operational efficiency, reduced processing time, and enhanced data accuracy. The system enables better decision-making through automated reporting and provides scalable foundation for future business growth.
## TECHNICAL FOUNDATION
Built on robust technical infrastructure supporting concurrent operations, error recovery, and system monitoring. The solution incorporates modern development practices with emphasis on reliability, performance, and maintainability.
"""
                
                with log_lock:
                    logger.info("Generated comprehensive project summary")
                    logger.debug(f"Summary length: {len(project_summary.split())} words")
                
                graphs = []
                try:
                    erd_graph = await generate_graph_from_requirement(project_summary, target_graph="entity relationship diagram")
                    if not erd_graph.generated_code.startswith("Error"):
                        graphs.append(erd_graph)
                    
                    req_graph = await generate_graph_from_requirement(project_summary, target_graph="requirement diagram")
                    if not req_graph.generated_code.startswith("Error"):
                        graphs.append(req_graph)
                    
                    with log_lock:
                        logger.info(f"Generated {len(graphs)} graphs for project summary")
                except Exception as e:
                    with log_lock:
                        logger.error(f"Error generating graphs for project summary: {e}")
                
                return project_summary, graphs

            except Exception as e:
                with log_lock:
                    logger.error(f"Error during LLM summary generation: {e}")
                return "Comprehensive summary generation failed due to processing error.", []
    
    except Exception as e:
        with log_lock:
            logger.error(f"Unexpected error in project summary generation: {e}")
        return "Summary generation encountered an unexpected issue.", []

async def generate_requirements_for_file_whole(file_path: str, base_dir: str, language: str) -> FileRequirements:
    """Generate requirements for a file with optimized RAG integration"""
    try:
        with log_lock:
            logger.debug(f"Processing file: {file_path}, Language: {language}")
        
        _, ext = os.path.splitext(file_path)
        is_excel = ext.lower() in {'.xlsx', '.xlsm', '.xls', '.xlsb'}
        
        if is_excel:
            language = "Excel" if ext.lower() == '.xlsx' else "VBA"
            content = await extract_excel_content(file_path)
        else:
            code_extensions = {'.py', '.js', '.jsx', '.ts', '.tsx', '.java', '.scala', '.rb', '.go', '.cpp', '.c', '.rs', '.kt', '.swift', '.php', '.cs'}
            if ext.lower() not in code_extensions:
                with log_lock:
                    logger.info(f"Skipping non-code file: {file_path} (Extension: {ext})")
                return FileRequirements(
                    relative_path=os.path.relpath(file_path, base_dir),
                    file_name=os.path.basename(file_path),
                    requirements="Non-code file: Requirements generation not applicable.",
                    dependencies=[]
                )
            
            # Read file content with improved error handling
            content = None
            try:
                async with aiofiles.open(file_path, 'r', encoding='utf-8') as f:
                    content = await f.read()
            except UnicodeDecodeError:
                # Try alternative encodings efficiently
                for encoding in ['latin-1', 'cp1252', 'iso-8859-1']:
                    try:
                        async with aiofiles.open(file_path, 'r', encoding=encoding) as f:
                            content = await f.read()
                        logger.warning(f"Successfully read {file_path} with {encoding} encoding")
                        break
                    except Exception:
                        continue
                
                # Final fallback to binary mode
                if content is None:
                    try:
                        async with aiofiles.open(file_path, 'rb') as f:
                            raw_content = await f.read()
                            content = raw_content.decode('latin-1', errors='replace')
                        logger.warning(f"Read {file_path} in binary mode with error replacement")
                    except Exception as e:
                        logger.error(f"Failed to read {file_path}: {str(e)}")
                        return _create_fallback_requirements(file_path, base_dir, language, "encoding error")
            except Exception as e:
                logger.error(f"Unexpected error reading {file_path}: {str(e)}")
                return _create_fallback_requirements(file_path, base_dir, language, "file read error")
        
        if content is None:
            return _create_fallback_requirements(file_path, base_dir, language, "could not decode file content")

        # Count tokens and determine processing strategy
        total_tokens = count_tokens(content)
        with log_lock:
            logger.debug(f"File: {file_path}, Language: {language}, Size: {len(content)} bytes, Tokens: {total_tokens}")
        
        # Process RAG storage in background (non-blocking)
        asyncio.create_task(_store_in_vector_db_async(content, file_path, language, base_dir))
        
        # Extract dependencies in parallel with requirements generation
        dependencies_task = asyncio.create_task(_extract_dependencies_async(file_path, content, language, base_dir))
        
        # Generate requirements based on file size
        if total_tokens < 1000:
            requirements_text = await _generate_small_file_requirements(content, file_path, language)
        elif total_tokens <= MAX_SAFE_TOKENS:
            requirements_text = await _generate_medium_file_requirements(content, file_path, language, total_tokens)
        else:
            requirements_text = await _generate_large_file_requirements(content, file_path, language, total_tokens)

        # Wait for dependencies to complete
        dependencies = await dependencies_task

        # Enrich each dependency with the Overview section from the requirements of the corresponding FileRequirements (if available in the current batch)
        import inspect
        frame = inspect.currentframe()
        all_requirements = None
        while frame:
            if 'all_requirements' in frame.f_locals:
                all_requirements = frame.f_locals['all_requirements']
                break
            frame = frame.f_back
        if all_requirements:
            for dep in dependencies:
                match = next((r for r in all_requirements if getattr(r, 'file_name', None) == dep.file_name), None)
                if match and getattr(match, 'requirements', None):
                    reqs = match.requirements.split('\n\n')
                    overview_section = next((s for s in reqs if s.strip().lower().startswith('overview')), None)
                    if overview_section:
                        lines = overview_section.split('\n')[1:]
                        overview = ' '.join(l.strip() for l in lines if l.strip())[:200]
                        dep.dependency_reason = overview

        # Add dependencies to requirements text
        dep_text = _format_dependencies(dependencies)
        requirements_text += f"\n\nDependent Files\n{dep_text}"

        # Ensure all dependencies are FileDependency model instances before passing to FileRequirements
        from app.services.dependency_service import FileDependency
        new_deps = []
        for dep in dependencies:
            if isinstance(dep, FileDependency):
                new_deps.append(dep)
            elif isinstance(dep, dict):
                new_deps.append(FileDependency(**dep))
            else:
                continue
        dependencies = new_deps

        # Return only the raw LLM output as requirements (no structured JSON)
        return FileRequirements(
            relative_path=os.path.relpath(file_path, base_dir),
            file_name=os.path.basename(file_path),
            requirements=requirements_text,
            dependencies=dependencies
        )
        
    except Exception as e:
        with log_lock:
            logger.error(f"Error generating requirements for {file_path}: {str(e)}")
        return FileRequirements(
            relative_path=os.path.relpath(file_path, base_dir),
            file_name=os.path.basename(file_path),
            requirements=f"Error generating requirements: {str(e)}",
            dependencies=[]
        )


async def _store_in_vector_db_async(content: str, file_path: str, language: str, base_dir: str):
    """Store content in vector database asynchronously without blocking main processing, with detailed logging and diagnostics."""
    try:
        from app.services.vector_store_service import PgVectorStoreService, VectorStoreService
        logger.info(f"[EMBED] Storing file: {file_path} (lang: {language})")
        if isinstance(vector_store, PgVectorStoreService):
            logger.info(f"[EMBED] Clearing previous embeddings for Postgres vector store.")
            vector_store.clear_embeddings()
        # Chunking
        if count_tokens(content) > 2000:
            chunks = await split_into_chunks(content, language, file_path)
        else:
            # Always emit at least one chunk
            chunks = [(content, 1, content.count('\n') + 1)]
        logger.info(f"[EMBED] File {file_path} split into {len(chunks)} chunk(s)")
        documents = []
        for i, (chunk, start_line, end_line) in enumerate(chunks):
            logger.info(f"[EMBED] Preparing chunk {i+1}/{len(chunks)} for {file_path}: lines {start_line}-{end_line}, tokens={count_tokens(chunk)}")
            documents.append({
                "content": chunk,
                "metadata": {
                    "file_path": file_path,
                    "language": language,
                    "start_line": start_line,
                    "end_line": end_line,
                    "base_dir": base_dir,
                    "chunk_index": i
                }
            })
        # Store in vector database
        try:
            await vector_store.add_documents(documents)
            logger.info(f"[EMBED] Successfully stored {len(documents)} chunk(s) for {file_path}")
        except Exception as embed_exc:
            logger.error(f"[EMBED] Error storing chunks for {file_path}: {embed_exc}")
        # Save only if using local FAISS vector store
        if isinstance(vector_store, VectorStoreService) and len(documents) > 10:
            try:
                vector_store.save(VECTOR_STORE_DIR)
                logger.info(f"[EMBED] Vector store saved after storing {file_path}")
            except Exception as save_exc:
                logger.error(f"[EMBED] Error saving vector store after {file_path}: {save_exc}")
    except Exception as e:
        logger.error(f"[EMBED] Failed to store chunks in vector store for {file_path}: {str(e)}")


async def _extract_dependencies_async(file_path: str, content: str, language: str, base_dir: str):
    """Extract dependencies asynchronously"""
    try:
        from app.services.dependency_service import detect_dependencies
        dependencies = await detect_dependencies(file_path, content, language, base_dir)
        
        validated_dependencies = []
        for dep in dependencies:
            try:
                validated_dep = FileDependency(**dep.dict())
                validated_dependencies.append(validated_dep)
            except ValidationError as ve:
                with log_lock:
                    logger.error(f"Invalid FileDependency for {file_path}: {dep.dict()}, Error: {str(ve)}")
        
        with log_lock:
            logger.debug(f"Validated dependencies for {file_path}: {[dep.dict() for dep in validated_dependencies]}")
        
        return validated_dependencies
    except Exception as e:
        logger.error(f"Error extracting dependencies for {file_path}: {str(e)}")
        return []


def _normalize_header(header):
    """Normalize section headers for flexible matching, including Markdown-style headers."""
    header = header.lower().strip()
    header = header.lstrip('#').strip()  # Remove leading Markdown hashes
    header = header.replace('functionality', 'functionalities')
    header = header.replace('key functionality', 'key functionalities')
    header = header.replace('dependent file', 'dependent files')
    header = header.replace('use-case', 'use case')
    return header

REQUIRED_SECTIONS = [
    'overview',
    'objective',
    'use case',
    'key functionalities',
    'workflow summary',
    'dependent files'
]

# Flexible section parsing
def _parse_sections(text):
    """Parse LLM output into sections, allowing for minor header variations and extra commentary."""
    sections = {}
    current = None
    # Updated regex to split on Markdown or plain section headers
    lines = re.split(r'\n+\s*#*\s*(?=Overview|Objective|Use Case|Key Functionalities|Workflow Summary|Dependent Files)', text, flags=re.IGNORECASE)
    for line in lines:
        norm = _normalize_header(line)
        for req in REQUIRED_SECTIONS:
            if norm.startswith(req):
                current = req
                sections[current] = []
                break
        else:
            if current:
                sections[current].append(line)
    # Join section lines
    for k in sections:
        sections[k] = '\n'.join(sections[k]).strip()
    return sections

# Update validation to accept partial/variant outputs

def _validate_and_format_requirements(req_text: str, file_path: str, language: str, rules_summary: str) -> str:
    """Validate and format LLM-generated requirements. Always return all required sections, filling with a file-specific explanation if missing."""
    sections = _parse_sections(req_text)
    formatted = []
    for req in REQUIRED_SECTIONS:
        content = sections.get(req, '').strip()
        # If LLM returned an error message, replace with explanation
        if content.startswith("Error: LLM failed") or content.startswith("No business logic detected") or not content or len(content) < 10:
            if req == 'dependent files':
                content = 'No dependencies detected.'
            elif req == 'overview' and rules_summary:
                content = rules_summary
            else:
                # File-specific fallback explanation
                content = (f"No explicit business logic was detected in the {os.path.basename(file_path)} {language} file for the '{req.title()}' section. "
                           f"This may be because the file primarily contains UI components, configuration, or utility code, or lacks business rules, workflows, or domain logic. "
                           f"If this is a React component or similar, the file may focus on rendering or wiring rather than business processes.")
        formatted.append(f"{req.title()}\n{content}\n")
    return '\n'.join(formatted)


# Add logging for prompt and output
async def _call_llm_with_logging(prompt, llm_func, file_path, attempt):
    logger = logging.getLogger("analysis_service")
    logger.debug(f"Prompt for {file_path} (attempt {attempt}):\n{prompt[:1000]}")
    output = await llm_func(prompt)
    logger.debug(f"Raw LLM output for {file_path} (attempt {attempt}):\n{output[:2000]}")
    return output

# Emergency prompt for stubborn files
EMERGENCY_PROMPT = """
You must output all five sections: Overview, Objective, Use Case, Key Functionalities, Workflow Summary, and Dependent Files. Each section must have at least 3 sentences. Reference every function, class, and comment in the file. Do not return generic or placeholder text. If you cannot find business logic, explain why for each section. Do not include any values or code, just pure business logic.
"""

# Special handling for small/script files
SMALL_FILE_PROMPT = """
This is a small or script-like file. Describe its purpose and logic in detail, referencing any operations, variables, or comments. Output all five sections as above.
"""

# Update requirements generation to use new logic
async def _generate_small_file_requirements(content: str, file_path: str, language: str) -> str:
    rules_detected = await detect_business_rules(content, language)
    rules_summary = (
        "\n".join([f"Detected {rule['domain']} rule ({rule['source']}): {rule['description']}" 
                  for rule in rules_detected])
        if rules_detected else ""
    )
    prompt = SMALL_FILE_PROMPT + "\n" + content
    # First attempt
    output1 = await _call_llm_with_logging(prompt, llm_config._llm, file_path, 1)
    reqs1 = _validate_and_format_requirements(output1, file_path, language, rules_summary)
    if not reqs1.startswith("Limited analysis"):
        return reqs1
    # Second attempt: more forceful
    prompt2 = prompt + "\nBe more detailed. Reference every function, class, and comment."
    output2 = await _call_llm_with_logging(prompt2, llm_config._llm, file_path, 2)
    reqs2 = _validate_and_format_requirements(output2, file_path, language, rules_summary)
    if not reqs2.startswith("Limited analysis"):
        return reqs2
    # Third attempt: emergency prompt
    prompt3 = EMERGENCY_PROMPT + "\n" + content

async def _generate_medium_file_requirements(content: str, file_path: str, language: str, total_tokens: int) -> str:
    """Hybrid RAG: Try direct LLM, fallback to RAG-augmented prompt for medium files"""
    from app.utils.llm_quality import is_generic_llm_output
    logger = logging.getLogger("analysis_service")
    if total_tokens > MAX_SAFE_TOKENS // 2:
        content = content[:int(len(content) * MAX_SAFE_TOKENS * 0.3 / total_tokens)]
        with log_lock:
            logger.warning(f"Truncated {file_path} to {len(content)} bytes")
    # 1. Fast path: Direct LLM
    prompt = STRICT_SECTION_PROMPT + "\n" + content
    output = await _call_llm_with_logging(prompt, llm_config._llm, file_path, 1, max_tokens=1000)
    logger.debug(f"[RAW LLM OUTPUT][{file_path}][direct]: {output[:1000]}")
    logger.info(f"[LLM OUTPUT] Chunk 1 (direct):\n{output}")
    if not is_generic_llm_output(output):
        logger.info(f"[HYBRID RAG] Used direct LLM output for {file_path}")
        return output
    # 2. Fallback: RAG-augmented
    try:
        rag_context_data = await vector_store.get_related_context(query=content, k=3)
        rag_context = rag_context_data.get("context", "")
    except Exception as e:
        logger.error(f"[HYBRID RAG] Retrieval failed for {file_path}: {e}")
        rag_context = ""
    rag_prompt = (
        STRICT_SECTION_PROMPT +
        f"\n\nHere is related context from the project:\n---\n{rag_context}\n---\n" +
        content
    )
    rag_output = await _call_llm_with_logging(rag_prompt, llm_config._llm, file_path, 2, max_tokens=1000)
    logger.debug(f"[RAW LLM OUTPUT][{file_path}][RAG]: {rag_output[:1000]}")
    logger.info(f"[LLM OUTPUT] Chunk 2 (RAG):\n{rag_output}")
    logger.info(f"[HYBRID RAG] Used RAG-augmented output for {file_path}")
    return rag_output

async def _generate_large_file_requirements(content: str, file_path: str, language: str, total_tokens: int) -> str:
    """Chunk large files and send each chunk to LLM; use hybrid RAG for each chunk; concatenate outputs"""
    from app.utils.llm_quality import is_generic_llm_output
    logger = logging.getLogger("analysis_service")
    with log_lock:
        logger.debug(f"Splitting {file_path} into chunks due to high token count: {total_tokens}")
    chunks = await split_into_chunks(content, language, file_path)
    with log_lock:
        logger.debug(f"Generated {len(chunks)} chunks for {file_path}")
    requirements_list = []
    for i, (chunk_content, start_line, end_line) in enumerate(chunks):
        # Hybrid logic per chunk
        prompt = STRICT_SECTION_PROMPT + "\n" + chunk_content
        output = await _call_llm_with_logging(prompt, llm_config._llm, f"{file_path}::chunk{i+1}", 1, max_tokens=1000)
        logger.debug(f"[DEBUG] LLM output for chunk {i+1} (direct):\n{output}")
        logger.info(f"[LLM OUTPUT] Chunk {i+1} (direct):\n{output}")
        if not is_generic_llm_output(output):
            logger.info(f"[HYBRID RAG][chunk {i+1}] Used direct LLM output for {file_path}")
            requirements_list.append(output)
            continue
        # Fallback: RAG-augmented
        try:
            rag_context_data = await vector_store.get_related_context(query=chunk_content, k=3)
            rag_context = rag_context_data.get("context", "")
        except Exception as e:
            logger.error(f"[HYBRID RAG][chunk {i+1}] Retrieval failed for {file_path}: {e}")
            rag_context = ""
        rag_prompt = (
            STRICT_SECTION_PROMPT +
            f"\n\nHere is related context from the project:\n---\n{rag_context}\n---\n" +
            chunk_content
        )
        rag_output = await _call_llm_with_logging(rag_prompt, llm_config._llm, f"{file_path}::chunk{i+1}", 2, max_tokens=1000)
        logger.debug(f"[DEBUG] LLM output for chunk {i+1} (RAG):\n{rag_output}")
        logger.info(f"[LLM OUTPUT] Chunk {i+1} (RAG):\n{rag_output}")
        logger.info(f"[HYBRID RAG][chunk {i+1}] Used RAG-augmented output for {file_path}")
        requirements_list.append(rag_output)
    # Instead of concatenating, combine and deduplicate sections
    return await combine_requirements(requirements_list, language)

def _create_fallback_requirements_text(file_path: str, language: str, error_type: str, rules_summary: str) -> str:
    """Create fallback requirements text"""
    return (
        f"Overview\nLimited analysis for {os.path.basename(file_path)} due to {error_type}. {rules_summary}\n\n"
        f"Objective\nTo support core {language.lower()} business functions.\n\n"
        f"Use Case\nGeneral {language.lower()} business processes.\n\n"
        f"Key Functionalities\n1. Basic Processing: Performs core tasks.\n\n2. Support Functions: Assists operations.\n\n"
        f"Workflow Summary\nIntegrates with {language.lower()} system workflow."
    )


def _format_dependencies(dependencies: list) -> str:
    """Format dependencies for display"""
    if not dependencies:
        return "No dependencies detected."
    
    dep_lines = []
    for dep in dependencies:
        dep_lines.append(dep.file_name)
        dep_lines.append(dep.dependency_reason.replace('\n', ' '))
    
    return "\\n".join(dep_lines)


def _create_chunk_fallback(chunk_num: int, file_path: str) -> str:
    """Create fallback text for failed chunk processing"""
    return (
        f"Overview\nFailed to process chunk {chunk_num} of {os.path.basename(file_path)}.\n\n"
        f"Objective\nTo support chunk functionality.\n\n"
        f"Use Case\nGeneral chunk operations.\n\n"
        f"Key Functionalities\n1. Partial Processing: Supports chunk {chunk_num} operations.\n\n2. Fallback: Provides basic functionality.\n\n"
        f"Workflow Summary\nProcesses available chunk data."
    )


async def generate_requirements_for_files_whole(directory: str) -> FilesRequirements:
    """Optimized batch processing of files with detailed summary logging for embedding/chunking."""
    logger.info(f"Starting requirements generation for {directory}")
    try:
        # Get all files to process
        code_files = get_code_files(directory)
        all_files = []
        for root, _, files in os.walk(directory):
            for file in files:
                file_path = os.path.join(root, file)
                _, ext = os.path.splitext(file)
                if ext.lower() in {'.xlsx', '.xlsm', '.xls', '.xlsb'}:
                    language = "Excel" if ext.lower() == '.xlsx' else "VBA"
                    all_files.append((file_path, language))
                elif (file_path, get_file_language(file_path)) in code_files:
                    all_files.append((file_path, get_file_language(file_path)))
        total_files = len(all_files)
        logger.info(f"Found {total_files} files for requirements generation in {directory}")
        files_requirements = []
        completed_count = 0
        chunking_stats = {"total_files": total_files, "total_chunks": 0, "files": []}
        batch_size = min(BATCH_SIZE, 6)
        file_batches = [all_files[i:i + batch_size] for i in range(0, total_files, batch_size)]
        for batch_num, batch in enumerate(file_batches):
            logger.info(f"Processing batch {batch_num + 1}/{len(file_batches)}")
            tasks = [
                generate_requirements_for_file_whole(file_path, directory, language)
                for file_path, language in batch
            ]
            batch_results = await asyncio.gather(*tasks, return_exceptions=True)
            for i, result in enumerate(batch_results):
                file_path, language = batch[i]
                if isinstance(result, Exception):
                    logger.error(f"Error processing {file_path}: {result}")
                    fallback = _create_fallback_requirements(file_path, directory, language, "processing error")
                    files_requirements.append(fallback)
                    chunking_stats["files"].append({"file": file_path, "chunks": 0, "status": "error", "reason": str(result)})
                else:
                    files_requirements.append(result)
                    try:
                        if hasattr(result, "requirements") and hasattr(result, "file_name"):
                            logger.info(f"[SUMMARY] File {result.file_name} processed.")
                        chunking_stats["files"].append({"file": file_path, "chunks": "?", "status": "ok"})
                    except Exception as stat_exc:
                        logger.warning(f"[SUMMARY] Could not inspect chunking for {file_path}: {stat_exc}")
            completed_count += len(batch)
            progress = completed_count / total_files * 100
            logger.info(f"Progress: {completed_count}/{total_files} files processed ({progress:.1f}%)")
            if batch_num < len(file_batches) - 1:
                await asyncio.sleep(0.1)
        # --- NEW: Build reverse dependency (dependents) map ---
        # Map: file_name -> list of FileRequirements that import it
        file_name_to_req = {req.file_name: req for req in files_requirements}
        dependents_map = {req.file_name: [] for req in files_requirements}
        for req in files_requirements:
            for dep in getattr(req, 'dependencies', []):
                dep_name = getattr(dep, 'file_name', None)
                if dep_name and dep_name in dependents_map:
                    # Add this file as a dependent of dep_name
                    overview = req.requirements.split('\n\n')[0] if req.requirements else ''
                    short_overview = overview.split('\n', 1)[-1][:200] if overview else ''
                    dependents_map[dep_name].append({
                        'file_name': req.file_name,
                        'relative_path': req.relative_path,
                        'dependency_reason': short_overview or 'Imports this file.'
                    })
        # Attach dependents to each FileRequirements
        for req in files_requirements:
            req.dependents = dependents_map.get(req.file_name, [])
        # Sort results by relative path
        files_requirements.sort(key=lambda x: x.relative_path)
        files_requirements_obj = FilesRequirements(files=files_requirements)
        project_summary, graphs = await generate_project_summary(files_requirements_obj)
        files_requirements_obj.project_summary = project_summary
        files_requirements_obj.graphs = graphs
        from app.services.vector_store_service import VectorStoreService
        try:
            if isinstance(vector_store, VectorStoreService):
                vector_store.save(VECTOR_STORE_DIR)
                logger.info("Vector store saved successfully")
        except Exception as e:
            logger.error(f"Failed to save vector store: {e}")
        logger.info(f"[SUMMARY] Embedding/chunking stats: {json.dumps(chunking_stats, indent=2)}")
        logger.info(f"Completed requirements generation for {len(files_requirements)} files with project summary and {len(graphs)} graphs")
        return files_requirements_obj
    except Exception as e:
        logger.error(f"Error processing directory {directory}: {str(e)}")
        return FilesRequirements(files=[])

def extract_json_from_llm_output(output_text):
    """Try to extract a JSON array from LLM output, even if it's inside a code block or not at the start."""
    import json, re
    # Try to find a code block with json
    match = re.search(r"```json\s*([\s\S]*?)\s*```", output_text)
    if match:
        json_str = match.group(1).strip()
        try:
            return json.loads(json_str)
        except Exception:
            pass
    # Try to find any code block
    match = re.search(r"```[\w]*\s*([\s\S]*?)\s*```", output_text)
    if match:
        json_str = match.group(1).strip()
        try:
            return json.loads(json_str)
        except Exception:
            pass
    # Try to find the first [ ... ]
    match = re.search(r"\[.*\]", output_text, re.DOTALL)
    if match:
        json_str = match.group(0)
        try:
            return json.loads(json_str)
        except Exception:
            pass
    # Try to parse the whole output
    try:
        return json.loads(output_text)
    except Exception:
        pass
    return None

# Update: Strict section enforcement, no default/fallback content, higher max_tokens, and forceful prompts
REQUIRED_SECTIONS = [
    'overview',
    'objective',
    'use case',
    'key functionalities',
    'workflow summary',
    'dependent files'
]

STRICT_SECTION_PROMPT = """
You must output all six sections: Overview, Objective, Use Case, Key Functionalities, Workflow Summary, and Dependent Files. Each section must be non-empty, file-specific, and reference every function, class, and comment in the file. Do not return generic or placeholder text. If you cannot find business logic, explain why for that section. Do not include any values or code, just pure business logic. Return only the six sections, no extra commentary.
"""

async def _call_llm_with_logging(prompt, llm_func, file_path, attempt, max_tokens=1000):
    logger = logging.getLogger("analysis_service")
    logger.debug(f"Prompt for {file_path} (attempt {attempt}):\n{prompt[:1000]}")
    output = await llm_func(prompt, max_tokens=max_tokens)
    logger.debug(f"Raw LLM output for {file_path} (attempt {attempt}):\n{output[:2000]}")
    return output

# Update requirements generation to use strict prompt and no fallback
async def _generate_small_file_requirements(content: str, file_path: str, language: str) -> str:
    rules_detected = await detect_business_rules(content, language)
    rules_summary = (
        "\n".join([f"Detected {rule['domain']} rule ({rule['source']}): {rule['description']}" 
                  for rule in rules_detected])
        if rules_detected else ""
    )
    prompt = STRICT_SECTION_PROMPT + f"\nDetected rules for context:\n{rules_summary}\n\n{content}"
    # First attempt
    output1 = await _call_llm_with_logging(prompt, llm_config._llm, file_path, 1, max_tokens=1000)
    reqs1 = _parse_sections(output1)
    if all(req in reqs1 and reqs1[req] and len(reqs1[req].strip()) > 10 for req in REQUIRED_SECTIONS):
        return '\n\n'.join([f"{req.title()}\n{reqs1[req].strip()}" for req in REQUIRED_SECTIONS])
    # Second attempt: more forceful
    prompt2 = prompt + "\nBe more detailed. Reference every function, class, and comment."
    output2 = await _call_llm_with_logging(prompt2, llm_config._llm, file_path, 2, max_tokens=1000)
    reqs2 = _parse_sections(output2)
    if all(req in reqs2 and reqs2[req] and len(reqs2[req].strip()) > 10 for req in REQUIRED_SECTIONS):
        return '\n\n'.join([f"{req.title()}\n{reqs2[req].strip()}" for req in REQUIRED_SECTIONS])
    # Third attempt: strictest prompt
    prompt3 = STRICT_SECTION_PROMPT + f"\n{content}"
    output3 = await _call_llm_with_logging(prompt3, llm_config._llm, file_path, 3, max_tokens=1000)
    reqs3 = _parse_sections(output3)
    if all(req in reqs3 and reqs3[req] and len(reqs3[req].strip()) > 10 for req in REQUIRED_SECTIONS):
        return '\n\n'.join([f"{req.title()}\n{reqs3[req].strip()}" for req in REQUIRED_SECTIONS])
    logger = logging.getLogger("analysis_service")
    logger.warning(f"LLM failed to generate all required sections for {file_path} after 3 attempts.")
    return f"Error: LLM failed to generate all required sections for {file_path}. Please check the file content or try again."

# File-type-specific example outputs
EXAMPLES = {
    'python': '''
Overview
This Python file implements a Flask-based REST API for managing employee records. It defines endpoints for creating, reading, updating, and deleting employee data, and integrates with a PostgreSQL database using SQLAlchemy. The file includes input validation, error handling, and logging for all operations.

Objective
To provide a secure and efficient API for CRUD operations on employee records, supporting integration with HR systems and internal dashboards.

Use Case
Used by HR staff and automated scripts to manage employee data, synchronize records with payroll systems, and generate reports for compliance and auditing.

Key Functionalities
1. Employee CRUD Operations: Endpoints for creating, reading, updating, and deleting employee records in the database.

2. Input Validation: Ensures all required fields are present and valid before processing requests.

3. Error Handling: Returns clear error messages and appropriate HTTP status codes for invalid operations.

4. Logging: Records all API requests and errors for auditing and debugging purposes.

Workflow Summary
The API receives HTTP requests, validates input, performs database operations, handles errors, and logs all activities. It supports integration with other business systems and ensures data integrity and security.

Dependent Files
No dependencies detected.''',
    'angular': '''
Overview
This Angular file defines a component for managing user profiles, including form validation, data binding, and integration with backend services. It uses Angular decorators to specify component metadata and leverages services for HTTP communication.

Objective
To provide a reusable and interactive UI component for user profile management, supporting both data entry and editing workflows.

Use Case
Used in the user management module to allow administrators and users to view and update profile information, with real-time validation and feedback.

Key Functionalities
1. Data Binding: Synchronizes form fields with component state and backend data.

2. Form Validation: Implements custom and built-in validators for user input.

3. Service Integration: Communicates with backend APIs to fetch and update user data.

4. UI Feedback: Provides real-time feedback and error messages to users.

Workflow Summary
The component initializes by loading user data, sets up form controls, handles user input, validates data, and submits changes to the backend. It updates the UI based on success or error responses.

Dependent Files
No dependencies detected.''',
    'csharp': '''
Overview
This C# file implements a service class for processing payroll data in a .NET application. It defines methods for calculating salaries, applying deductions, and generating payroll reports.

Objective
To automate payroll processing and ensure accurate salary calculations in compliance with company policies.

Use Case
Used by HR and finance teams to process employee payroll, generate payslips, and maintain payroll records.

Key Functionalities
1. Salary Calculation: Computes gross and net salaries based on employee data and company rules.

2. Deduction Handling: Applies tax, insurance, and other deductions as per regulations.

3. Report Generation: Produces payroll reports for management and compliance.

Workflow Summary
The service retrieves employee data, performs calculations, applies deductions, and generates reports. It integrates with other modules for data consistency.

Dependent Files
No dependencies detected.''',
    'sql': '''
Overview
This SQL file defines stored procedures and views for managing sales transactions in the database. It includes logic for inserting, updating, and querying sales records.

Objective
To centralize and optimize sales data management, ensuring data integrity and efficient reporting.

Use Case
Used by sales and analytics teams to record transactions, update sales data, and generate sales performance reports.

Key Functionalities
1. Insert Procedure: Adds new sales records to the database.

2. Update Procedure: Modifies existing sales records based on business rules.

3. Reporting Views: Provides aggregated sales data for analysis.

Workflow Summary
The procedures and views interact with sales tables, enforce business rules, and support reporting needs. They are invoked by application logic and scheduled jobs.

Dependent Files
No dependencies detected.''',
    'html': '''
Overview
This HTML file defines the structure and layout for the user dashboard page, including navigation, content sections, and embedded widgets. It uses semantic HTML tags and integrates with CSS and JavaScript for styling and interactivity.

Objective
To provide a user-friendly and accessible dashboard interface for end users, supporting navigation and data visualization.

Use Case
Used as the main entry point for users to access dashboard features, view key metrics, and interact with widgets.

Key Functionalities
1. Navigation Bar: Provides links to main sections of the application.

2. Content Sections: Organizes dashboard widgets and information panels.

3. Widget Embedding: Integrates charts and tables for data visualization.

Workflow Summary
The HTML structure is loaded by the browser, styled by CSS, and enhanced by JavaScript. It supports dynamic updates and user interactions.

Dependent Files
No dependencies detected.''',
    'excel': '''
Overview
This Excel macro-enabled file automates the calculation of project costs and generates summary reports. It includes VBA macros for data validation, cost adjustments, and summary table generation across multiple worksheets.

Objective
To streamline project cost calculations and reporting by automating data entry, validation, and summary generation using macros.

Use Case
Used by project managers and finance teams to quickly calculate costs, apply adjustments, and produce summary reports for internal review and external audits.

Key Functionalities
1. Macro-Driven Cost Calculation: Automates the calculation of project costs using VBA macros, including adjustments for tax, discount, and overhead.

2. Data Validation: Checks for missing or invalid data before calculations proceed, ensuring data quality.

3. Summary Table Generation: Compiles results from multiple worksheets into a summary table for reporting.

Workflow Summary
The macro processes input data, validates entries, performs calculations, and generates a summary report. It handles errors and missing data, ensuring accurate and reliable outputs.

Dependent Files
No dependencies detected.''',
    'config': '''
Overview
This JSON configuration file defines application settings, environment variables, and dependency references for the project. It is used to control runtime behavior and manage external integrations.

Objective
To centralize configuration management, making it easy to update settings and dependencies without modifying code.

Use Case
Used by developers and DevOps teams to manage environment-specific settings, API keys, and service endpoints.

Key Functionalities
1. Environment Settings: Defines variables for different deployment environments (dev, staging, prod).

2. Dependency References: Lists external services and libraries required by the application.

3. Feature Flags: Enables or disables features based on configuration.

Workflow Summary
The application loads this configuration at startup, applying settings and initializing dependencies as specified. Changes to this file affect application behavior without code changes.

Dependent Files
No dependencies detected.'''
}

# Helper to get file type for prompt selection
FILE_TYPE_MAP = {
    '.py': 'python', '.js': 'python', '.jsx': 'python', '.ts': 'angular', '.tsx': 'angular', '.java': 'python', '.cs': 'csharp',
    '.sql': 'sql', '.html': 'html',
    '.xlsm': 'excel', '.xlsx': 'excel', '.xls': 'excel',
    '.json': 'config', '.csproj': 'config', '.sln': 'config', '.yml': 'config', '.yaml': 'config', '.http': 'config'
}

def get_prompt_example(file_path):
    _, ext = os.path.splitext(file_path)
    return EXAMPLES.get(FILE_TYPE_MAP.get(ext.lower(), 'python'), EXAMPLES['python'])

async def _generate_requirements_with_section_completion(content: str, file_path: str, language: str, rules_summary: str) -> str:
    # Compose prompt as before, but return LLM output as-is, no fallback/validation
    _, ext = os.path.splitext(file_path)
    technical_context = ''
    if ext.lower() in ['.ts', '.html']:
        technical_context = summarize_angular_file(content, file_path)
    elif ext.lower() in ['.js', '.jsx', '.tsx']:
        from app.utils.angular_parser import summarize_angular_file
        technical_context = summarize_angular_file(content, file_path)
    elif ext.lower() in ['.cs']:
        from app.utils.dotnet_parser import summarize_csharp_code
        technical_context = summarize_csharp_code(content)
    elif ext.lower() in ['.csproj', '.sln', '.json', '.yml', '.yaml', '.http']:
        from app.utils.dotnet_parser import summarize_dotnet_config
        technical_context = summarize_dotnet_config(content, file_path)
    elif ext.lower() == '.sql':
        try:
            from app.utils.sql_parser import summarize_sql_file
            technical_context = summarize_sql_file(content, file_path)
        except ImportError:
            technical_context = 'SQL file; summary utility not implemented.'
    example_output = get_prompt_example(file_path)
    prompt_base = f"You are a senior business analyst. Carefully review the following {language} file and produce a highly detailed, file-specific business requirements document.\nAnalyze all functions, classes, comments, and business logic. For each section below, provide a content-rich, file-specific explanation based strictly on the actual code and comments.\n- Do NOT use generic, vague, or placeholder statements.\n- If the file is small, expand on every available detail, inferring business purpose, logic, and workflow from names, comments, and code structure.\n- Use the following detected rules for context:\n{rules_summary}\n- Technical context: {technical_context}\n- Your output MUST have six sections: Overview, Objective, Use Case, Key Functionalities, Workflow Summary, Dependent Files, separated by two newlines.\n- Each section must be at least 2-3 sentences, and Key Functionalities must be a numbered list with a blank line between each.\n- Example output:\n{example_output}\n- Return ONLY the six sections, no extra commentary.\n\nAnalyze this file:\n```{language}\n{content}\n```"
    max_tokens = 1000
    logger = logging.getLogger("analysis_service")
    try:
        response = await _call_llm_with_logging(prompt_base, llm_config._llm, file_path, 1, max_tokens)
        logger.debug(f"[RAW LLM OUTPUT][{file_path}]: {response[:1000]}")
        return response
    except Exception as e:
        logger.error(f"[LLM ERROR] Exception during LLM requirements for {file_path}: {e}")
        return ""
# Use this in all requirements generation functions
async def _generate_small_file_requirements(content: str, file_path: str, language: str) -> str:
    from app.utils.llm_quality import is_generic_llm_output
    logger = logging.getLogger("analysis_service")
    # 1. Fast path: Direct LLM
    prompt = STRICT_SECTION_PROMPT + "\n" + content
    output = await _call_llm_with_logging(prompt, llm_config._llm, file_path, 1, max_tokens=1000)
    logger.debug(f"[RAW LLM OUTPUT][{file_path}][direct]: {output[:1000]}")
    logger.info(f"[LLM OUTPUT] Chunk 1 (direct):\n{output}")
    if not is_generic_llm_output(output):
        logger.info(f"[HYBRID RAG] Used direct LLM output for {file_path}")
        return output
    # 2. Fallback: RAG-augmented
    try:
        rag_context_data = await vector_store.get_related_context(query=content, k=3)
        rag_context = rag_context_data.get("context", "")
    except Exception as e:
        logger.error(f"[HYBRID RAG] Retrieval failed for {file_path}: {e}")
        rag_context = ""
    rag_prompt = (
        STRICT_SECTION_PROMPT +
        f"\n\nHere is related context from the project:\n---\n{rag_context}\n---\n" +
        content
    )
    rag_output = await _call_llm_with_logging(rag_prompt, llm_config._llm, file_path, 2, max_tokens=1000)
    logger.debug(f"[RAW LLM OUTPUT][{file_path}][RAG]: {rag_output[:1000]}")
    logger.info(f"[LLM OUTPUT] Chunk 2 (RAG):\n{rag_output}")
    logger.info(f"[HYBRID RAG] Used RAG-augmented output for {file_path}")
    return rag_output